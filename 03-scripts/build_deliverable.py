"""Assemble the final deliverable xlsx.

Reads:
  - 01-inputs/vendors-raw.xlsx (preserves original structure)
  - 04-outputs/reconciled.jsonl
  - 04-outputs/top3.json
  - 04-outputs/executive-memo.md (if it exists; otherwise leaves the tab empty)

Writes:
  - 04-outputs/Vendor Analysis Assessment - Ojas Shah.xlsx

Optional flag:
  --memo-url <url>    URL of the published Google Doc memo. When set, the
                      Recommendations tab is populated with a clickable link
                      to the Doc PLUS the full memo body underneath, so a
                      reviewer can either click through or read inline.

The /publish-to-drive slash command runs this script twice: once without
--memo-url (to produce an xlsx for the initial upload), then again with
--memo-url after the Google Doc is created, and re-uploads the xlsx.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from openpyxl.styles import Alignment, Font

from _common import INPUTS_DIR, OUTPUTS_DIR, read_jsonl

OUT_NAME = "Vendor Analysis Assessment - Ojas Shah.xlsx"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--memo-url",
        default="",
        help="Google Doc URL to prepend to the Recommendations tab.",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(INPUTS_DIR / "vendors-raw.xlsx")

    reconciled = {row["name"]: row for row in read_jsonl(OUTPUTS_DIR / "reconciled.jsonl")}

    # ---------- Vendor Analysis sheet ----------
    ws = wb["Vendor Analysis Assessment"]
    for row in ws.iter_rows(min_row=2):
        name_cell = row[0]
        if not name_cell.value:
            continue
        match = reconciled.get(str(name_cell.value).strip())
        if not match:
            continue
        ws.cell(row=name_cell.row, column=2, value=match.get("department", ""))
        ws.cell(row=name_cell.row, column=4, value=match.get("description", ""))
        ws.cell(row=name_cell.row, column=5, value=match.get("recommendation", ""))

    # ---------- Top 3 Opportunities ----------
    top3_path = OUTPUTS_DIR / "top3.json"
    if top3_path.exists():
        top3 = json.loads(top3_path.read_text(encoding="utf-8"))
        ws3 = wb["Top 3 Opportunities"]
        for i, opp in enumerate(top3.get("opportunities", [])):
            row_idx = 2 + i
            ws3.cell(row=row_idx, column=1, value=f"Opportunity {opp['rank']}")
            ws3.cell(row=row_idx, column=2, value=opp["title"])
            explanation = (
                f"{opp['implementation_summary']}\n\n"
                f"Vendors involved: {', '.join(opp['vendors_involved'])}\n"
                f"Current annual spend: ${opp['current_annual_spend']:,.0f}\n"
                f"Projected savings: ${opp['projected_annual_savings']:,.0f} "
                f"({opp['projected_savings_pct']*100:.0f}%)\n"
                f"Implementation effort: {opp['implementation_effort']}\n"
                f"Key risk: {opp['risk_notes']}"
            )
            ws3.cell(row=row_idx, column=3, value=explanation)
            ws3.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")

        total = top3.get("total_projected_savings", 0)
        ws3.cell(row=6, column=1, value="Total")
        ws3.cell(row=6, column=2, value=f"${total:,.0f}/year identified")
        ws3.cell(row=6, column=1).font = Font(bold=True)
        ws3.cell(row=6, column=2).font = Font(bold=True)

    # ---------- Methodology ----------
    ws_m = wb["Methodology"]
    ws_m.cell(row=2, column=1, value=_build_methodology_text())
    ws_m.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws_m.row_dimensions[2].height = 400

    # ---------- Recommendations (executive memo) ----------
    memo_path = OUTPUTS_DIR / "executive-memo.md"
    if memo_path.exists():
        memo_body = memo_path.read_text(encoding="utf-8")
        if args.memo_url:
            content = (
                f"Google Doc version: {args.memo_url}\n\n"
                f"(Memo body also included below for inline review.)\n\n"
                f"{'-' * 60}\n\n"
                f"{memo_body}"
            )
        else:
            content = memo_body
        ws_r = wb["CEOCFO Recommendations"]
        ws_r.cell(row=2, column=1, value=content)
        ws_r.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws_r.row_dimensions[2].height = 500

    out_path = OUTPUTS_DIR / OUT_NAME
    wb.save(out_path)
    print(f"Wrote {out_path}")
    if args.memo_url:
        print(f"  Memo URL embedded: {args.memo_url}")


def _build_methodology_text() -> str:
    return (
        "METHODOLOGY\n\n"
        "1. SETUP. Built a reproducible Claude Code project (CLAUDE.md, .claude/agents, "
        ".claude/commands, versioned prompts in 02-prompts/, scripts in 03-scripts/). "
        "Every prompt used in the analysis is checked into the project so the work is auditable. "
        "Source data was fetched directly from Google Drive via the Anthropic Google Drive MCP "
        "using the /setup-source slash command.\n\n"
        "2. FIRST-PASS ENRICHMENT. The 386 vendors were processed in batches of 25 by "
        "analyze_vendors.py, calling Claude with the prompt at 02-prompts/01-first-pass-prompt.md. "
        "Each call received the full batch (so peer-vendor context was available for "
        "Consolidate detection) and returned structured JSON with department, description, "
        "recommendation, reasoning, and a self-reported confidence score.\n\n"
        "3. INDEPENDENT QC PASS. qc_validate.py re-classified every vendor using a "
        "different prompt (02-prompts/02-qc-pass-prompt.md) that biases toward conservatism. "
        "Crucially, the QC pass did NOT see the first pass's output — independence is the "
        "whole point of validation.\n\n"
        "4. RECONCILIATION. reconcile.py compared the two passes line by line. Where they "
        "agreed, the answer was accepted (with averaged confidence). Where they disagreed, "
        "a tiebreaker prompt (03-reconciliation-prompt.md) saw both reasonings and made the "
        "final call. The full disagreement log is in 05-validation/qc-report.md — every "
        "vendor that needed a tiebreaker is listed there with both analyses and the "
        "resolution reasoning.\n\n"
        "5. STRATEGIC SYNTHESIS. top_opportunities.py rolled vendors up by department and "
        "function, then prompted Claude with the spend-weighted view to identify the three "
        "highest-impact opportunities. Output is constrained to be specific (named vendors), "
        "financially justified (current spend → projected savings %), and realistic (effort + "
        "risk assessment).\n\n"
        "6. EXECUTIVE MEMO. The /exec-memo slash command produced a 1-page CFO/CEO memo "
        "as markdown. The /publish-to-drive slash command then uploaded it to Google Drive "
        "as a native Google Doc (Drive auto-converted from text/plain).\n\n"
        "7. PUBLISH. The /publish-to-drive slash command also uploaded this completed "
        "spreadsheet to the user's Google Drive via the Drive MCP. The Recommendations tab "
        "contains a link to the published Google Doc memo.\n\n"
        "TOOLS USED: Claude Code CLI, Anthropic Python SDK (claude-sonnet-4-5 model), "
        "Anthropic Google Drive MCP (read source + publish deliverables), openpyxl.\n\n"
        "PROMPTS: All prompts live as versioned markdown files in 02-prompts/. The full "
        "project is available at the GitHub repo provided with this submission."
    )


if __name__ == "__main__":
    main()
